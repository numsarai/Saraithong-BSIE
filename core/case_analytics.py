"""
case_analytics.py
-----------------
Deterministic case-level analytics for bulk statement runs.
"""

from __future__ import annotations

import json
from collections import defaultdict

from openpyxl.styles import Font, PatternFill
from pathlib import Path
from typing import Any

import pandas as pd


def _clean_text(value: object) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in {"", "nan", "none", "nat"} else text


def _clean_number(value: object) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _identity_label(account: str, name: str) -> str:
    if account and name:
        return f"{name} ({account})"
    return account or name or "UNKNOWN"


def _normalize_counterparty_id(row: pd.Series) -> str:
    account = _clean_text(row.get("counterparty_account"))
    partial = _clean_text(row.get("partial_account"))
    name = _clean_text(row.get("counterparty_name"))
    if account:
        return account
    if partial:
        return f"PARTIAL:{partial}"
    if name:
        return f"NAME:{name}"
    return "UNKNOWN"


def _normalize_counterparty_label(row: pd.Series) -> str:
    account = _clean_text(row.get("counterparty_account"))
    partial = _clean_text(row.get("partial_account"))
    name = _clean_text(row.get("counterparty_name"))
    if account and name:
        return f"{name} ({account})"
    if partial and name:
        return f"{name} ({partial})"
    return name or account or partial or "UNKNOWN"


def _build_components(edges: list[tuple[str, str]]) -> list[list[str]]:
    graph: dict[str, set[str]] = defaultdict(set)
    for left, right in edges:
        if not left or not right or left == "UNKNOWN" or right == "UNKNOWN":
            continue
        graph[left].add(right)
        graph[right].add(left)

    components: list[list[str]] = []
    visited: set[str] = set()
    for node in sorted(graph):
        if node in visited:
            continue
        stack = [node]
        component: list[str] = []
        while stack:
            current = stack.pop()
            if current in visited:
                continue
            visited.add(current)
            component.append(current)
            stack.extend(sorted(graph[current] - visited))
        components.append(sorted(component))
    return components


def compute_case_analytics(summary: dict[str, Any]) -> dict[str, Any]:
    files = summary.get("files", []) or []
    processed_files = [row for row in files if row.get("status") == "processed"]

    counterparties: dict[str, dict[str, Any]] = {}
    fan_in: dict[str, dict[str, Any]] = defaultdict(lambda: {"count": 0, "value": 0.0, "counterparties": set()})
    fan_out: dict[str, dict[str, Any]] = defaultdict(lambda: {"count": 0, "value": 0.0, "counterparties": set()})
    bridge_accounts: dict[str, dict[str, Any]] = defaultdict(lambda: {"subject_accounts": set(), "labels": set(), "transactions": 0, "total_value": 0.0})
    graph_edges: list[tuple[str, str]] = []
    flagged_accounts: list[dict[str, Any]] = []

    for row in processed_files:
        account = _clean_text(row.get("account"))
        subject_name = _clean_text(row.get("name"))
        output_dir = Path(str(row.get("output_dir") or ""))
        if not account or not output_dir.exists():
            continue

        if (
            bool(row.get("bank_ambiguous"))
            or _clean_text(row.get("reconciliation_status")) in {"FAILED", "PARTIAL", "INFERRED"}
            or float(row.get("bank_confidence") or 0.0) < 0.75
        ):
            reasons = []
            if bool(row.get("bank_ambiguous")):
                reasons.append("ambiguous_bank_detection")
            if float(row.get("bank_confidence") or 0.0) < 0.75:
                reasons.append("low_bank_confidence")
            recon_status = _clean_text(row.get("reconciliation_status"))
            if recon_status in {"FAILED", "PARTIAL", "INFERRED"}:
                reasons.append(f"reconciliation_{recon_status.lower()}")
            flagged_accounts.append(
                {
                    "account": account,
                    "name": subject_name,
                    "bank": _clean_text(row.get("bank_name") or row.get("bank_key")),
                    "bank_confidence": round(float(row.get("bank_confidence") or 0.0), 4),
                    "reconciliation_status": recon_status,
                    "reason_codes": reasons,
                }
            )

        txn_path = output_dir / "processed" / "transactions.csv"
        if not txn_path.exists():
            continue

        tx_df = pd.read_csv(txn_path, dtype=str, encoding="utf-8-sig").fillna("")
        if tx_df.empty:
            continue
        tx_df["amount"] = tx_df["amount"].map(_clean_number)

        for _, tx in tx_df.iterrows():
            cp_id = _normalize_counterparty_id(tx)
            cp_label = _normalize_counterparty_label(tx)
            amount_abs = abs(_clean_number(tx.get("amount")))
            direction = _clean_text(tx.get("direction"))
            txn_date = _clean_text(tx.get("date"))

            cp_bucket = counterparties.setdefault(
                cp_id,
                {
                    "counterparty_id": cp_id,
                    "counterparty_label": cp_label,
                    "transaction_count": 0,
                    "total_value": 0.0,
                    "subject_accounts": set(),
                    "first_seen": txn_date,
                    "last_seen": txn_date,
                },
            )
            cp_bucket["counterparty_label"] = cp_label or cp_bucket["counterparty_label"]
            cp_bucket["transaction_count"] += 1
            cp_bucket["total_value"] += amount_abs
            cp_bucket["subject_accounts"].add(account)
            if txn_date:
                if not cp_bucket["first_seen"] or txn_date < cp_bucket["first_seen"]:
                    cp_bucket["first_seen"] = txn_date
                if not cp_bucket["last_seen"] or txn_date > cp_bucket["last_seen"]:
                    cp_bucket["last_seen"] = txn_date

            if cp_id not in {"", "UNKNOWN"}:
                bridge_accounts[cp_id]["subject_accounts"].add(account)
                bridge_accounts[cp_id]["labels"].add(cp_label)
                bridge_accounts[cp_id]["transactions"] += 1
                bridge_accounts[cp_id]["total_value"] += amount_abs
                graph_edges.append((account, cp_id))

            if direction == "IN":
                fan_in[account]["count"] += 1
                fan_in[account]["value"] += amount_abs
                fan_in[account]["counterparties"].add(cp_id)
            elif direction == "OUT":
                fan_out[account]["count"] += 1
                fan_out[account]["value"] += amount_abs
                fan_out[account]["counterparties"].add(cp_id)

    top_by_count = sorted(
        (
            {
                "counterparty_id": key,
                "counterparty_label": value["counterparty_label"],
                "transaction_count": int(value["transaction_count"]),
                "total_value": round(float(value["total_value"]), 2),
                "subject_accounts": sorted(value["subject_accounts"]),
                "first_seen": value["first_seen"] or "",
                "last_seen": value["last_seen"] or "",
            }
            for key, value in counterparties.items()
            if key != "UNKNOWN"
        ),
        key=lambda item: (-item["transaction_count"], -item["total_value"], item["counterparty_label"]),
    )[:10]

    top_by_value = sorted(
        top_by_count,
        key=lambda item: (-item["total_value"], -item["transaction_count"], item["counterparty_label"]),
    )[:10]

    def build_rank(source: dict[str, dict[str, Any]], metric_name: str) -> list[dict[str, Any]]:
        return sorted(
            (
                {
                    "subject_account": account,
                    "subject_label": _identity_label(
                        account,
                        next((_clean_text(row.get("name")) for row in processed_files if _clean_text(row.get("account")) == account), ""),
                    ),
                    "transaction_count": int(values["count"]),
                    "total_value": round(float(values["value"]), 2),
                    "unique_counterparties": len(values["counterparties"]),
                    "metric": metric_name,
                }
                for account, values in source.items()
            ),
            key=lambda item: (-item["transaction_count"], -item["total_value"], item["subject_account"]),
        )

    bridge_rows = sorted(
        (
            {
                "counterparty_id": cp_id,
                "counterparty_label": sorted(data["labels"])[0] if data["labels"] else cp_id,
                "subject_accounts": sorted(data["subject_accounts"]),
                "subject_account_count": len(data["subject_accounts"]),
                "transaction_count": int(data["transactions"]),
                "total_value": round(float(data["total_value"]), 2),
            }
            for cp_id, data in bridge_accounts.items()
            if len(data["subject_accounts"]) > 1
        ),
        key=lambda item: (-item["subject_account_count"], -item["transaction_count"], -item["total_value"], item["counterparty_label"]),
    )

    connected_groups = [
        {
            "group_id": f"GRP-{index + 1:03d}",
            "size": len(component),
            "nodes": component,
        }
        for index, component in enumerate(sorted(_build_components(graph_edges), key=lambda item: (-len(item), item)))
    ]

    overview = {
        "processed_accounts": len(processed_files),
        "flagged_accounts": len(flagged_accounts),
        "connected_groups": len(connected_groups),
        "bridge_accounts": len(bridge_rows),
        "largest_counterparty_by_count": top_by_count[0] if top_by_count else None,
        "largest_counterparty_by_value": top_by_value[0] if top_by_value else None,
    }

    return {
        "run_id": summary.get("run_id", ""),
        "overview": overview,
        "top_counterparties_by_count": top_by_count,
        "top_counterparties_by_value": top_by_value,
        "fan_in_rank": build_rank(fan_in, "fan_in"),
        "fan_out_rank": build_rank(fan_out, "fan_out"),
        "bridge_accounts": bridge_rows,
        "connected_groups": connected_groups,
        "flagged_accounts": sorted(flagged_accounts, key=lambda item: item["account"]),
    }


def write_case_analytics(run_dir: Path, analytics: dict[str, Any]) -> None:
    run_dir.mkdir(parents=True, exist_ok=True)
    (run_dir / "case_analytics.json").write_text(
        json.dumps(analytics, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    with pd.ExcelWriter(run_dir / "case_analytics.xlsx", engine="openpyxl") as writer:
        pd.DataFrame([analytics.get("overview", {})]).to_excel(writer, sheet_name="Overview", index=False)
        pd.DataFrame(analytics.get("top_counterparties_by_count", [])).to_excel(writer, sheet_name="Top_By_Count", index=False)
        pd.DataFrame(analytics.get("top_counterparties_by_value", [])).to_excel(writer, sheet_name="Top_By_Value", index=False)
        pd.DataFrame(analytics.get("fan_in_rank", [])).to_excel(writer, sheet_name="Fan_In", index=False)
        pd.DataFrame(analytics.get("fan_out_rank", [])).to_excel(writer, sheet_name="Fan_Out", index=False)
        pd.DataFrame(analytics.get("bridge_accounts", [])).to_excel(writer, sheet_name="Bridge_Accounts", index=False)
        pd.DataFrame(analytics.get("connected_groups", [])).to_excel(writer, sheet_name="Connected_Groups", index=False)
        pd.DataFrame(analytics.get("flagged_accounts", [])).to_excel(writer, sheet_name="Flagged_Accounts", index=False)

        # Apply TH Sarabun New font to all sheets
        from core.exporter import REPORT_FONT_NAME, REPORT_FONT_SIZE
        data_font = Font(name=REPORT_FONT_NAME, size=REPORT_FONT_SIZE)
        hdr_font = Font(name=REPORT_FONT_NAME, size=REPORT_FONT_SIZE, bold=True, color="FFFFFF")
        hdr_fill = PatternFill(fill_type="solid", fgColor="1E293B")
        for ws in writer.book.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if cell.row == 1:
                        cell.font = hdr_font
                        cell.fill = hdr_fill
                    else:
                        cell.font = data_font
