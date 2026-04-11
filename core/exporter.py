"""
exporter.py
-----------
Exports the Account Package to disk in CSV and multi-sheet Excel (.xlsx).

Output structure:
  /data/output/{account_number}/
      raw/original.xlsx
      processed/
          {name}_{bank}_report.xlsx    ← multi-sheet report
          transactions.csv
          entities.csv   entities.xlsx
          links.csv
      meta.json
"""

import json
import logging
import re
import shutil
from pathlib import Path
from typing import Optional, Union, Tuple

import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

from core.account_parser import parse_account
from utils.date_utils import format_date_range
from core.export_anx import export_anx_from_graph
from core.export_i2_import import write_i2_import_package
from core.graph_analysis import build_graph_analysis, write_graph_analysis_exports
from core.graph_export import write_graph_exports
from core.bank_logo_registry import build_bank_logo_catalog, find_bank_logo_record, render_bank_logo_png_bytes
from core.ofx_io import export_ofx
from core.reconciliation import reconcile_balances
from project_meta import APP_CONTACT_PHONE, APP_DEVELOPER_NAME, APP_OWNER_NAME

logger = logging.getLogger(__name__)

from paths import BUILTIN_CONFIG_DIR, CONFIG_DIR, OUTPUT_DIR as BASE_OUTPUT


TRANSACTION_EXPORT_COLUMNS = [
    "transaction_id",
    "date",
    "time",
    "transaction_type",
    "direction",
    "amount",
    "currency",
    "classification_source",
    "classification_reason",
    "classification_review_flag",
    "classification_model",
    "heuristic_transaction_type",
    "heuristic_confidence",
    "ai_transaction_type",
    "ai_confidence",
    "ai_counterparty_name",
    "balance",
    "balance_source",
    "expected_balance",
    "balance_difference",
    "balance_check_status",
    "subject_account",
    "subject_name",
    "counterparty_account",
    "partial_account",
    "counterparty_name",
    "from_account",
    "to_account",
    "bank",
    "channel",
    "description",
    "confidence",
    "is_overridden",
    "override_from_account",
    "override_to_account",
    "override_reason",
    "override_by",
    "override_timestamp",
    "raw_account_value",
    "parsed_account_type",
    "nlp_type_hint",
    "nlp_confidence",
    "nlp_best_name",
    "source_file",
    "row_number",
]

ENTITY_EXPORT_COLUMNS = [
    "entity_id",
    "entity_type",
    "entity_label",
    "identity_value",
    "account_number",
    "name",
    "first_seen",
    "last_seen",
    "transaction_count",
    "source_transaction_ids",
]

LINK_EXPORT_COLUMNS = [
    "link_id",
    "transaction_id",
    "date",
    "transaction_type",
    "direction",
    "amount",
    "currency",
    "from_account",
    "from_entity_id",
    "from_label",
    "to_account",
    "to_entity_id",
    "to_label",
    "counterparty_name",
    "description",
]

TRANSACTION_CATEGORY_RULES = {
    "transfer_in": {"transaction_type": "IN_TRANSFER"},
    "transfer_out": {"transaction_type": "OUT_TRANSFER"},
    "deposit": {"transaction_type": "DEPOSIT"},
    "withdraw": {"transaction_type": "WITHDRAW"},
}

SUMMARY_SHEET_SPECS = [
    ("ready to transfer in", "transfer_in"),
    ("ready to transfer out", "transfer_out"),
    ("ready to deposit", "deposit"),
    ("ready to withdraw", "withdraw"),
]

SUMMARY_SHEET_COLUMNS = [
    "Subject Account Name",
    "Subject Account Number",
    "Transaction Date Time",
    "Transaction Amount",
    "Counterparty Account Name",
    "Counterparty Account Number",
]

WORKBOOK_SHEET_NAMES = {
    "cover": "Report_Cover",
    "all_transactions": "All_Transactions",
    "transfer_in_detail": "Transfer_In",
    "transfer_out_detail": "Transfer_Out",
    "deposit_detail": "Deposits",
    "withdraw_detail": "Withdrawals",
    "entities": "Entities",
    "links": "Links",
    "reconciliation": "Reconciliation",
    "bank_logos": "Bank_Logos",
}
LOGO_SHEET_HEADERS = ["Logo", "Bank Name", "Key", "Template Status", "Category", "Notes"]


def _safe_filename(name: str) -> str:
    """Sanitise a string for use in a file name (keep Thai, alphanumeric, space, dash)."""
    s = re.sub(r'[\\/:*?"<>|]', '', name).strip()
    # collapse multiple spaces / dots
    s = re.sub(r'\s+', ' ', s)
    return s or "unknown"


def _cleanup_stale_report_files(processed_dir: Path, bank: str, keep_filename: str) -> None:
    """Remove older report workbooks for the same bank so reruns keep one canonical file."""
    keep_path = processed_dir / keep_filename
    bank_suffix = f"_{_safe_filename(bank)}_report.xlsx"
    for candidate in processed_dir.iterdir():
        if not candidate.is_file() or candidate.suffix.lower() != ".xlsx":
            continue
        if candidate == keep_path:
            continue
        if candidate.name == "report.xlsx" or candidate.name.endswith(bank_suffix):
            try:
                candidate.unlink()
                logger.info("  Removed stale report file: %s", candidate.name)
            except Exception as exc:
                logger.warning("  Could not remove stale report file %s: %s", candidate.name, exc)


def _ensure_dirs(account_number: str) -> Tuple[Path, Path, Path]:
    """
    Create output directory structure for an account.
    Returns (account_dir, raw_dir, processed_dir).
    """
    account_dir  = BASE_OUTPUT / account_number
    raw_dir      = account_dir / "raw"
    processed_dir = account_dir / "processed"

    raw_dir.mkdir(parents=True, exist_ok=True)
    processed_dir.mkdir(parents=True, exist_ok=True)

    return account_dir, raw_dir, processed_dir


def _write_csv_and_excel(df: pd.DataFrame, processed_dir: Path, stem: str) -> None:
    """Write a DataFrame as both CSV and Excel."""
    csv_path   = processed_dir / f"{stem}.csv"
    excel_path = processed_dir / f"{stem}.xlsx"

    df.to_csv(csv_path,   index=False, encoding="utf-8-sig")   # utf-8-sig for Excel Thai compat
    df.to_excel(excel_path, index=False, engine="openpyxl")

    logger.info(f"  Exported: {csv_path.name} + {excel_path.name}")


def _reorder_columns(df: pd.DataFrame, preferred_columns: list[str]) -> pd.DataFrame:
    """Return a copy with preferred columns first and any extras preserved."""
    available = [c for c in preferred_columns if c in df.columns]
    extras = [c for c in df.columns if c not in available]
    return df[available + extras].copy()


def _format_export_date(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    for splitter in ("T", " "):
        if splitter in text:
            text = text.split(splitter, 1)[0]
            break
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        year, month, day = text.split("-")
        return f"{day} {month} {year}"
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}", text):
        day, month, year = text.split("/")
        return f"{day.zfill(2)} {month.zfill(2)} {year}"
    return text


def _format_export_amount(value: object, *, absolute: bool = False) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        amount = float(text.replace(",", ""))
    except ValueError:
        return text
    if absolute:
        amount = abs(amount)
    if float(amount).is_integer():
        return f"{int(amount):,}"
    return f"{amount:,.2f}"


def _stringify_report_account(value: object) -> str:
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return ""
    parsed = parse_account(value)
    if parsed.get("clean"):
        return str(parsed["clean"])
    digits_only = re.sub(r"\D", "", text)
    return digits_only or text


def _combine_report_datetime(date_value: object, time_value: object) -> str:
    date_text = _format_export_date(date_value)
    time_text = str(time_value or "").strip()
    if date_text and time_text:
        return f"{date_text} {time_text}"
    return date_text or time_text


def _resolve_report_counterparty_account(row: pd.Series) -> str:
    counterparty_account = _stringify_report_account(row.get("counterparty_account", ""))
    if counterparty_account:
        return counterparty_account
    return _stringify_report_account(row.get("partial_account", ""))


def _prepare_summary_sheet(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

    rows = []
    for _, row in df.iterrows():
        rows.append({
            "Subject Account Name": str(row.get("subject_name", "") or "").strip(),
            "Subject Account Number": _stringify_report_account(row.get("subject_account", "")),
            "Transaction Date Time": _combine_report_datetime(row.get("date", ""), row.get("time", "")),
            "Transaction Amount": _format_export_amount(row.get("amount", ""), absolute=True),
            "Counterparty Account Name": str(row.get("counterparty_name", "") or "").strip(),
            "Counterparty Account Number": _resolve_report_counterparty_account(row),
        })

    return pd.DataFrame(rows, columns=SUMMARY_SHEET_COLUMNS)


def _build_summary_sheet_map(transaction_categories: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    return {
        sheet_name: _prepare_summary_sheet(transaction_categories.get(category_key, pd.DataFrame()))
        for sheet_name, category_key in SUMMARY_SHEET_SPECS
    }

def _format_transaction_frame_for_export(df: pd.DataFrame, *, absolute_amount: bool = True) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    formatted = df.copy()
    if "date" in formatted.columns:
        formatted["date"] = formatted["date"].map(_format_export_date)
    for money_col in ("amount", "balance", "expected_balance", "balance_difference"):
        if money_col in formatted.columns:
            formatted[money_col] = formatted[money_col].map(
                lambda value, col=money_col: _format_export_amount(value, absolute=absolute_amount if col == "amount" else False)
            )
    return formatted


def _format_links_frame_for_export(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    formatted = df.copy()
    if "date" in formatted.columns:
        formatted["date"] = formatted["date"].map(_format_export_date)
    if "amount" in formatted.columns:
        formatted["amount"] = formatted["amount"].map(lambda value: _format_export_amount(value, absolute=True))
    return formatted


def _prepare_transaction_exports(transactions: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    """Build the master transaction table plus deterministic category splits."""
    tx_all = _reorder_columns(transactions, TRANSACTION_EXPORT_COLUMNS).reset_index(drop=True)
    category_frames: dict[str, pd.DataFrame] = {}
    for export_key, rule in TRANSACTION_CATEGORY_RULES.items():
        filtered = tx_all[tx_all["transaction_type"] == rule["transaction_type"]].copy()
        category_frames[export_key] = filtered.reset_index(drop=True)
    return tx_all, category_frames


def _make_entity_label(row: pd.Series) -> str:
    """Return a human-readable entity label for CSV/Excel export."""
    name = str(row.get("name", "") or "").strip()
    account = str(row.get("account_number", "") or "").strip()
    if name and account:
        return f"{name} ({account})"
    return name or account or "UNKNOWN"


def _prepare_entities_export(entities: pd.DataFrame) -> pd.DataFrame:
    """Add i2-friendly identity and label columns to the entity export."""
    if entities.empty:
        return pd.DataFrame(columns=ENTITY_EXPORT_COLUMNS)

    prepared = entities.copy()
    prepared["entity_label"] = prepared.apply(_make_entity_label, axis=1)
    prepared["identity_value"] = prepared.apply(
        lambda row: str(row.get("account_number", "") or "").strip()
        or str(row.get("name", "") or "").strip()
        or "UNKNOWN",
        axis=1,
    )
    return _reorder_columns(prepared, ENTITY_EXPORT_COLUMNS).reset_index(drop=True)


def _resolve_entity_lookup_key(value: str) -> str:
    """Normalize link account identifiers into the entity-table lookup form."""
    value = str(value or "").strip()
    if value.startswith("PARTIAL:"):
        return value.replace("PARTIAL:", "", 1)
    return value or "UNKNOWN"


def _prepare_links_export(links: pd.DataFrame, entities: pd.DataFrame, transactions: pd.DataFrame) -> pd.DataFrame:
    """Add entity references and labels to links for direct graph import use."""
    if links.empty:
        return pd.DataFrame(columns=LINK_EXPORT_COLUMNS)

    entity_by_identity: dict[str, dict] = {}
    for _, row in entities.iterrows():
        identity = str(row.get("identity_value", "") or "").strip()
        if identity:
            entity_by_identity[identity] = row.to_dict()

    txn_details = transactions[[
        c for c in ["transaction_id", "direction", "currency", "counterparty_name", "description"]
        if c in transactions.columns
    ]].copy()
    merged = links.merge(txn_details, how="left", on="transaction_id")

    def map_entity(value: object) -> tuple[str, str]:
        identity = _resolve_entity_lookup_key(str(value or ""))
        entity = entity_by_identity.get(identity, {})
        entity_id = str(entity.get("entity_id", "") or "")
        label = str(entity.get("entity_label", "") or identity or "UNKNOWN")
        return entity_id, label

    from_refs = merged["from_account"].map(map_entity)
    to_refs = merged["to_account"].map(map_entity)

    prepared = merged.copy()
    prepared["link_id"] = prepared.get("transaction_id", "").map(lambda v: f"LNK-{v}" if v else "")
    prepared["from_entity_id"] = [entity_id for entity_id, _ in from_refs]
    prepared["from_label"] = [label for _, label in from_refs]
    prepared["to_entity_id"] = [entity_id for entity_id, _ in to_refs]
    prepared["to_label"] = [label for _, label in to_refs]

    return _reorder_columns(prepared, LINK_EXPORT_COLUMNS).reset_index(drop=True)


def _write_transactions_multisheet(
    transactions: pd.DataFrame,
    transaction_categories: dict[str, pd.DataFrame],
    transaction_summaries: dict[str, pd.DataFrame],
    entities: pd.DataFrame,
    links: pd.DataFrame,
    reconciliation: pd.DataFrame,
    bank_logo_catalog: list[dict],
    processed_dir: Path,
    report_filename: str = "report.xlsx",
    highlight_bank_key: str = "",
    report_context: dict | None = None,
) -> Path:
    """Write transactions Excel with i2-friendly category sheets + entities + links.

    Returns the Path to the written Excel file.
    """
    excel_path = processed_dir / report_filename

    sheet_map = {
        **transaction_summaries,
        WORKBOOK_SHEET_NAMES["all_transactions"]: transactions,
        WORKBOOK_SHEET_NAMES["transfer_in_detail"]: transaction_categories["transfer_in"],
        WORKBOOK_SHEET_NAMES["transfer_out_detail"]: transaction_categories["transfer_out"],
        WORKBOOK_SHEET_NAMES["deposit_detail"]: transaction_categories["deposit"],
        WORKBOOK_SHEET_NAMES["withdraw_detail"]: transaction_categories["withdraw"],
        WORKBOOK_SHEET_NAMES["entities"]: entities,
        WORKBOOK_SHEET_NAMES["links"]: links,
        WORKBOOK_SHEET_NAMES["reconciliation"]: reconciliation,
    }

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        cover_sheet = writer.book.create_sheet(WORKBOOK_SHEET_NAMES["cover"], 0)
        cover_sheet.sheet_view.showGridLines = False
        context = report_context or {}
        cover_sheet["B2"] = "BSIE Report Workbook"
        cover_sheet["B2"].font = Font(size=20, bold=True, color="1E293B")
        cover_sheet["B3"] = "Bank Statement Intelligence Engine"
        cover_sheet["B3"].font = Font(size=11, color="475569")
        source_bank_name = str(context.get("bank_name") or "")
        source_subject_name = str(context.get("subject_name") or "") or "Unknown subject"
        cover_items = [
            ("Subject Account Name", source_subject_name),
            ("Subject Account Number", str(context.get("account_number") or "")),
            ("Bank", source_bank_name),
            ("Original File", str(context.get("original_filename") or "")),
            ("Date Range", str(context.get("date_range") or "")),
            ("Transaction Count", str(context.get("transaction_count") or 0)),
            ("Program Owner / Developer", APP_OWNER_NAME or APP_DEVELOPER_NAME),
            ("Contact", APP_CONTACT_PHONE),
        ]
        for index, (label, value) in enumerate(cover_items, start=5):
            cover_sheet[f"B{index}"] = label
            cover_sheet[f"B{index}"].font = Font(size=10, bold=True, color="64748B")
            cover_sheet[f"C{index}"] = value
            cover_sheet[f"C{index}"].font = Font(size=11, color="0F172A")
        cover_sheet["B13"] = "Report Notes"
        cover_sheet["B13"].font = Font(size=11, bold=True, color="1E293B")
        cover_sheet["B14"] = "This workbook keeps analysis-ready views while preserving the original source file separately in the raw evidence folder."
        cover_sheet["B14"].alignment = Alignment(wrap_text=True, vertical="top")
        cover_sheet["B15"] = "Sheets 2-5 provide ready-to-use transfer in, transfer out, deposit, and withdraw views from normalized subject and counterparty data."
        cover_sheet["B15"].alignment = Alignment(wrap_text=True, vertical="top")
        cover_sheet["B16"] = "The final Bank_Logos sheet is a reference for current bank templates and prepared Thai bank logo assets."
        cover_sheet["B16"].alignment = Alignment(wrap_text=True, vertical="top")
        cover_sheet.column_dimensions["A"].width = 4
        cover_sheet.column_dimensions["B"].width = 18
        cover_sheet.column_dimensions["C"].width = 48
        cover_sheet.row_dimensions[2].height = 28
        cover_sheet.row_dimensions[14].height = 34
        cover_sheet.row_dimensions[15].height = 34
        cover_sheet.row_dimensions[16].height = 34

        cover_key = highlight_bank_key or str(find_bank_logo_record(display_name=source_bank_name).get("key") or "")
        cover_logo = XLImage(
            render_bank_logo_png_bytes(
                cover_key,
                display_name=source_bank_name or source_subject_name,
                has_template=True if cover_key else None,
                size=(144, 144),
            )
        )
        cover_logo.width = 92
        cover_logo.height = 92
        cover_sheet.add_image(cover_logo, "E2")

        for sheet_name, df in sheet_map.items():
            df_clean = df.reset_index(drop=True)
            df_clean.to_excel(writer, sheet_name=sheet_name, index=False)

            # Auto-fit column widths
            ws = writer.sheets[sheet_name]
            for col_cells in ws.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in col_cells if cell.value is not None),
                    default=8,
                )
                col_letter = col_cells[0].column_letter
                ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        logo_sheet = writer.book.create_sheet(WORKBOOK_SHEET_NAMES["bank_logos"])
        header_fill = PatternFill(fill_type="solid", fgColor="1E293B")
        header_font = Font(color="FFFFFF", bold=True)
        highlight_fill = PatternFill(fill_type="solid", fgColor="DBEAFE")
        logo_sheet.freeze_panes = "B2"
        logo_sheet.append(LOGO_SHEET_HEADERS)
        for cell in logo_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_index, bank in enumerate(bank_logo_catalog, start=2):
            logo_sheet.row_dimensions[row_index].height = 56
            logo_payload = render_bank_logo_png_bytes(
                bank.get("key"),
                display_name=str(bank.get("name") or ""),
                has_template=bool(bank.get("has_template")),
                size=(72, 72),
            )
            image = XLImage(logo_payload)
            image.width = 40
            image.height = 40
            logo_sheet.add_image(image, f"A{row_index}")
            logo_sheet.cell(row=row_index, column=2, value=str(bank.get("name") or ""))
            logo_sheet.cell(row=row_index, column=3, value=str(bank.get("key") or ""))
            logo_sheet.cell(row=row_index, column=4, value=str(bank.get("template_badge") or ""))
            logo_sheet.cell(row=row_index, column=5, value=str(bank.get("bank_type") or ""))
            note_parts = []
            if str(bank.get("template_source") or "") == "custom":
                note_parts.append("Custom template")
            elif str(bank.get("template_source") or "") == "builtin":
                note_parts.append("Built-in template")
            if str(bank.get("template_status") or "") == "logo_ready":
                note_parts.append("Prepared for future template work")
            if highlight_bank_key and str(bank.get("key") or "") == highlight_bank_key:
                note_parts.append("Report source bank")
            logo_sheet.cell(row=row_index, column=6, value="; ".join(note_parts))
            if highlight_bank_key and str(bank.get("key") or "") == highlight_bank_key:
                for col in range(1, 7):
                    logo_sheet.cell(row=row_index, column=col).fill = highlight_fill

        for col_letter, width in {"A": 12, "B": 34, "C": 18, "D": 26, "E": 18, "F": 34}.items():
            logo_sheet.column_dimensions[col_letter].width = width

    logger.info(f"  Multi-sheet report: {excel_path.name} ({len(sheet_map) + 2} sheets)")
    return excel_path


def export_package(
    transactions: pd.DataFrame,
    entities: pd.DataFrame,
    links: pd.DataFrame,
    account_number: str,
    bank: str,
    original_file: Union[str, Path],
    bank_key: str = "",
    subject_name: str = "",
    job_id: Optional[str] = None,
    reconciliation_df: Optional[pd.DataFrame] = None,
    reconciliation_summary: Optional[dict] = None,
) -> Path:
    """
    Export the full Account Package to disk.

    Parameters
    ----------
    transactions   : processed transaction DataFrame
    entities       : entity DataFrame
    links          : graph links DataFrame
    account_number : subject account number (used as folder name)
    bank           : bank name string
    original_file  : path to original Excel input file
    subject_name   : account holder name (used in report filename)

    Returns
    -------
    Path – path to account output directory
    """
    account_dir, raw_dir, processed_dir = _ensure_dirs(account_number)

    # 1. Copy original file
    orig_path = Path(original_file)
    dest_orig = raw_dir / f"original{orig_path.suffix.lower() or '.dat'}"
    try:
        shutil.copy2(orig_path, dest_orig)
        logger.info(f"Copied original: {dest_orig}")
    except Exception as e:
        logger.warning(f"Could not copy original file: {e}")

    # 2. Export transactions CSV (for programmatic access)
    transactions_export, transaction_categories = _prepare_transaction_exports(transactions)
    entities_export = _prepare_entities_export(entities)
    links_export = _prepare_links_export(links, entities_export, transactions_export)
    if reconciliation_df is None or reconciliation_summary is None:
        reconciliation_result = reconcile_balances(transactions_export)
        reconciliation_export = reconciliation_result.reconciliation
        reconciliation_summary = reconciliation_result.summary
        transactions_export = reconciliation_result.transactions
        transaction_categories = {
            export_key: transactions_export[transactions_export["transaction_type"] == rule["transaction_type"]].copy().reset_index(drop=True)
            for export_key, rule in TRANSACTION_CATEGORY_RULES.items()
        }
    else:
        reconciliation_export = reconciliation_df.copy()

    transactions_display = _format_transaction_frame_for_export(transactions_export, absolute_amount=True)
    transaction_categories_display = {
        export_key: _format_transaction_frame_for_export(df, absolute_amount=True)
        for export_key, df in transaction_categories.items()
    }
    transaction_summary_sheets = _build_summary_sheet_map(transaction_categories)
    links_display = _format_links_frame_for_export(links_export)
    reconciliation_display = _format_transaction_frame_for_export(reconciliation_export, absolute_amount=False)

    transactions_display.to_csv(processed_dir / "transactions.csv", index=False, encoding="utf-8-sig")
    logger.info("  Exported: transactions.csv")

    for export_key, df in transaction_categories_display.items():
        _write_csv_and_excel(df, processed_dir, export_key)

    ofx_path = processed_dir / "account.ofx"
    ofx_path.write_text(
        export_ofx(transactions_export, account_number=account_number, bank=bank, subject_name=subject_name),
        encoding="utf-8",
    )
    logger.info("  Exported: account.ofx")

    # 3. Export entities CSV + Excel
    _write_csv_and_excel(entities_export, processed_dir, "entities")

    # 4. Export links CSV
    _write_csv_and_excel(links_display, processed_dir, "links")

    # 4a. Export graph CSVs + manifest using the shared deterministic schema
    graph_batch_identity = f"LEGACY:{account_number}"
    graph_batch_label = f"{bank or 'Statement'} {account_number}".strip()
    graph_bundle = write_graph_exports(
        processed_dir,
        transactions=transactions_export,
        batch_identity=graph_batch_identity,
        batch_label=graph_batch_label,
    )
    logger.info("  Exported: nodes/edges/aggregated/derived graph files in CSV + JSON + graph_manifest.json")

    graph_analysis = build_graph_analysis(
        transactions_export,
        batch_identity=graph_batch_identity,
        batch_label=graph_batch_label,
        graph_bundle=graph_bundle,
    )
    graph_analysis_paths = write_graph_analysis_exports(processed_dir, graph_analysis)
    logger.info("  Exported: graph_analysis.json + graph_analysis.xlsx + suspicious_findings.csv + suspicious_findings.json")

    # 4a. Export reconciliation CSV + Excel
    _write_csv_and_excel(reconciliation_display, processed_dir, "reconciliation")

    # 4b. Export i2 Analyst's Notebook XML (.anx)
    anx_path = processed_dir / "i2_chart.anx"
    try:
        export_anx_from_graph(graph_bundle["nodes_df"], graph_bundle["edges_df"], anx_path)
    except Exception as e:
        logger.warning(f"  ANX export failed (non-fatal): {e}")

    # 4c. Export i2 import package (.ximp + companion CSV)
    i2_import_paths = write_i2_import_package(
        graph_bundle["nodes_df"],
        graph_bundle["edges_df"],
        processed_dir,
        subject=f"BSIE import for {subject_name or account_number or 'statement'}",
        comments=(
            f"Generated from BSIE account package for {subject_name or account_number or 'statement'} "
            "using the shared graph export bundle."
        ),
        author="BSIE",
    )

    # 5. Multi-sheet report Excel (all sheets in one file)
    #    Filename: {subject_name}_{bank}_report.xlsx  (or report.xlsx if no name)
    if subject_name:
        report_name = f"{_safe_filename(subject_name)}_{_safe_filename(bank)}_report.xlsx"
    else:
        report_name = "report.xlsx"
    _cleanup_stale_report_files(processed_dir, bank, report_name)
    template_entries = []
    for config_dir, template_source in ((BUILTIN_CONFIG_DIR, "builtin"), (CONFIG_DIR, "custom")):
        if not config_dir.exists():
            continue
        for config_path in sorted(config_dir.glob("*.json")):
            try:
                cfg = json.loads(config_path.read_text(encoding="utf-8"))
                template_entries.append({
                    "key": config_path.stem,
                    "name": cfg.get("bank_name", config_path.stem.upper()),
                    "template_source": template_source,
                })
            except Exception:
                logger.debug("Skipping malformed bank config in logo catalog: %s", config_path)
    logo_catalog = build_bank_logo_catalog(template_entries)
    highlight_bank = str(bank_key or "").strip().lower()
    if not highlight_bank and bank:
        highlight_bank = str(find_bank_logo_record(display_name=bank).get("key") or "").strip().lower()
    report_path = _write_transactions_multisheet(
        transactions_display,
        transaction_categories_display,
        transaction_summary_sheets,
        entities_export,
        links_display,
        reconciliation_display,
        logo_catalog,
        processed_dir,
        report_filename=report_name,
        highlight_bank_key=highlight_bank,
        report_context={
            "subject_name": subject_name,
            "account_number": account_number,
            "bank_name": bank,
            "original_filename": orig_path.name,
            "date_range": format_date_range(transactions_export["date"].tolist()) if "date" in transactions_export.columns else "",
            "transaction_count": len(transactions_export),
        },
    )

    # 7. Build and write meta.json
    meta = _build_meta(transactions, account_number, bank)
    meta["original_filename"] = orig_path.name
    meta["report_filename"] = report_path.name
    meta["reconciliation"] = reconciliation_summary or {}
    meta["graph_manifest"] = graph_bundle["manifest"]
    meta["graph_analysis"] = graph_analysis
    meta["category_files"] = {
        "all_transactions": "transactions.csv",
        "transfer_in": "transfer_in.csv",
        "transfer_out": "transfer_out.csv",
        "deposit": "deposit.csv",
        "withdraw": "withdraw.csv",
        "entities": "entities.csv",
        "links": "links.csv",
        "nodes": "nodes.csv",
        "nodes_json": "nodes.json",
        "edges": "edges.csv",
        "edges_json": "edges.json",
        "aggregated_edges": "aggregated_edges.csv",
        "aggregated_edges_json": "aggregated_edges.json",
        "derived_account_edges": "derived_account_edges.csv",
        "derived_account_edges_json": "derived_account_edges.json",
        "graph_manifest": "graph_manifest.json",
        "graph_analysis": graph_analysis_paths["json_path"].name,
        "graph_analysis_workbook": graph_analysis_paths["xlsx_path"].name,
        "suspicious_findings": graph_analysis_paths["suspicious_csv_path"].name,
        "suspicious_findings_json": graph_analysis_paths["suspicious_json_path"].name,
        "reconciliation": "reconciliation.csv",
        "ofx": "account.ofx",
        "anx": "i2_chart.anx",
        "i2_import_csv": i2_import_paths["csv_path"].name,
        "i2_import_spec": i2_import_paths["spec_path"].name,
    }
    meta["category_counts"] = {
        "transfer_in": len(transaction_categories["transfer_in"]),
        "transfer_out": len(transaction_categories["transfer_out"]),
        "deposit": len(transaction_categories["deposit"]),
        "withdraw": len(transaction_categories["withdraw"]),
        "graph_nodes": int(graph_bundle["manifest"].get("node_count", 0)),
        "graph_edges": int(graph_bundle["manifest"].get("edge_count", 0)),
        "graph_aggregated_edges": int(graph_bundle["manifest"].get("aggregated_edge_count", 0)),
        "graph_derived_account_edges": int(graph_bundle["manifest"].get("derived_account_edge_count", 0)),
        "graph_review_candidates": int(graph_analysis.get("overview", {}).get("review_candidate_nodes", 0)),
    }
    meta_path = account_dir / "meta.json"
    with meta_path.open("w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2, default=str)
    logger.info(f"  Meta written: {meta_path}")

    # Write JobMeta row to DB (non-fatal if DB unavailable)
    _write_job_meta_to_db(job_id, account_number, meta)

    logger.info(f"Account package complete: {account_dir}")
    return account_dir


def _write_job_meta_to_db(job_id: Optional[str], account_number: str, meta: dict) -> None:
    """Insert a JobMeta row into the SQLite DB. No-op if job_id is None or DB fails."""
    if not job_id:
        return
    try:
        from database import get_session, JobMeta
        from datetime import datetime as _dt
        with get_session() as session:
            row = JobMeta(
                account_number=account_number,
                job_id=job_id,
                bank=meta.get("bank", ""),
                total_in=float(meta.get("total_in", 0.0)),
                total_out=float(meta.get("total_out", 0.0)),
                total_circulation=float(meta.get("total_circulation", 0.0)),
                num_transactions=int(meta.get("num_transactions", 0)),
                date_range=meta.get("date_range", ""),
                num_unknown=int(meta.get("num_unknown", 0)),
                num_partial_accounts=int(meta.get("num_partial_accounts", 0)),
                report_filename=meta.get("report_filename", ""),
                created_at=_dt.utcnow(),
            )
            session.add(row)
            session.commit()
        logger.info(f"  JobMeta written to DB for job_id={job_id}")
    except Exception as e:
        logger.warning(f"  Could not write JobMeta to DB (non-fatal): {e}")


def _build_meta(transactions: pd.DataFrame, account_number: str, bank: str) -> dict:
    """Compute summary statistics for meta.json."""
    if transactions.empty:
        return {
            "account_number": account_number,
            "bank": bank,
            "total_in": 0,
            "total_out": 0,
            "num_transactions": 0,
            "date_range": "",
            "num_unknown": 0,
            "num_partial_accounts": 0,
        }

    amounts  = transactions["amount"].fillna(0)
    total_in  = float(amounts[amounts > 0].sum())
    total_out = float(amounts[amounts < 0].sum())

    raw_dates = transactions["date"].dropna().tolist()
    date_range = format_date_range(raw_dates) if raw_dates else ""

    if "counterparty_account" in transactions.columns:
        num_unknown = int((transactions["counterparty_account"] == "").sum())
    else:
        num_unknown = 0
    if "partial_account" in transactions.columns:
        num_partial = int((transactions["partial_account"] != "").sum())
    else:
        num_partial = 0

    total_circulation = round(total_in + abs(total_out), 2)

    return {
        "account_number":      account_number,
        "bank":                bank,
        "total_in":            round(total_in, 2),
        "total_out":           round(total_out, 2),
        "total_circulation":   total_circulation,
        "num_transactions":    len(transactions),
        "date_range":          date_range,
        "num_unknown":         num_unknown,
        "num_partial_accounts": num_partial,
    }
