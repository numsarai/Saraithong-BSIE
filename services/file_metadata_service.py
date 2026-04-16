"""
file_metadata_service.py
------------------------
Extract and verify file metadata for forensic integrity checks.
Detects: file tampering, creation/modification dates, authoring software.
Inspired by computer-forensics and metadata-extraction skills.
"""
from __future__ import annotations

import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


def _resolve_safe_path(
    file_path: str | Path,
    allowed_bases: list[Path],
) -> Path | None:
    """Resolve a file path and validate it stays under one of the allowed bases.

    Returns the resolved path on success, ``None`` if the path escapes
    every allowed base. The resolved path is the only handle callers should
    use downstream — the original ``file_path`` input must be considered
    untrusted data.
    """
    resolved = Path(file_path).resolve()  # codeql[py/path-injection]
    for base in allowed_bases:
        base_resolved = base.resolve()
        if resolved == base_resolved or base_resolved in resolved.parents:
            return resolved
    return None


def extract_file_metadata(file_path: str | Path) -> dict[str, Any]:
    """Extract metadata from a file for forensic verification."""
    from paths import INPUT_DIR, EVIDENCE_DIR, OUTPUT_DIR, EXPORTS_DIR

    allowed_dirs = [INPUT_DIR, EVIDENCE_DIR, OUTPUT_DIR, EXPORTS_DIR]
    safe_path = _resolve_safe_path(file_path, allowed_dirs)

    if safe_path is None:
        return {"error": "Access denied — path outside allowed directories"}

    # safe_path is provably under an allowed base at this point.
    if not safe_path.is_file():  # codeql[py/path-injection]
        return {"error": "File not found"}

    stat = safe_path.stat()  # codeql[py/path-injection]
    result: dict[str, Any] = {
        "path": str(safe_path),
        "filename": safe_path.name,
        "extension": safe_path.suffix.lower(),
        "file_size_bytes": stat.st_size,
        "file_size_display": _format_size(stat.st_size),
        "created_at": datetime.fromtimestamp(stat.st_birthtime).isoformat() if hasattr(stat, "st_birthtime") else None,
        "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        "accessed_at": datetime.fromtimestamp(stat.st_atime).isoformat(),
        "sha256": _compute_sha256(safe_path),
    }

    # Excel-specific metadata
    if safe_path.suffix.lower() in (".xlsx", ".xls"):
        result["excel_metadata"] = _extract_excel_metadata(safe_path)

    # Forensic checks
    result["integrity_checks"] = _run_integrity_checks(result)

    return result


def _compute_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _format_size(size_bytes: int) -> str:
    if size_bytes >= 1_048_576:
        return f"{size_bytes / 1_048_576:.1f} MB"
    if size_bytes >= 1024:
        return f"{size_bytes / 1024:.1f} KB"
    return f"{size_bytes} B"


def _extract_excel_metadata(path: Path) -> dict[str, Any]:
    """Extract workbook properties from Excel files."""
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        props = wb.properties
        meta = {
            "title": props.title or "",
            "subject": props.subject or "",
            "creator": props.creator or "",
            "last_modified_by": props.lastModifiedBy or "",
            "created": props.created.isoformat() if props.created else None,
            "modified": props.modified.isoformat() if props.modified else None,
            "description": props.description or "",
            "category": props.category or "",
            "keywords": props.keywords or "",
            "revision": props.revision or "",
            "sheet_names": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
        }
        wb.close()
        return meta
    except Exception as e:
        return {"error": str(e)}


def _run_integrity_checks(meta: dict[str, Any]) -> list[dict[str, Any]]:
    """Run forensic integrity checks on the file metadata."""
    checks: list[dict[str, Any]] = []

    # Check 1: File modification date vs creation date
    created = meta.get("created_at")
    modified = meta.get("modified_at")
    if created and modified:
        created_dt = datetime.fromisoformat(created)
        modified_dt = datetime.fromisoformat(modified)
        delta_hours = (modified_dt - created_dt).total_seconds() / 3600
        if delta_hours > 24 * 30:  # Modified more than 30 days after creation
            checks.append({
                "check": "file_modification_gap",
                "severity": "medium",
                "message": f"File modified {delta_hours / 24:.0f} days after creation — may indicate editing",
                "created": created,
                "modified": modified,
            })
        else:
            checks.append({
                "check": "file_modification_gap",
                "severity": "ok",
                "message": "File modification date is within normal range",
            })

    # Check 2: Excel metadata — last modified by different person
    excel_meta = meta.get("excel_metadata", {})
    if isinstance(excel_meta, dict) and not excel_meta.get("error"):
        creator = (excel_meta.get("creator") or "").strip().lower()
        modifier = (excel_meta.get("last_modified_by") or "").strip().lower()
        if creator and modifier and creator != modifier:
            checks.append({
                "check": "different_modifier",
                "severity": "warning",
                "message": f"Created by '{excel_meta['creator']}' but last modified by '{excel_meta['last_modified_by']}'",
                "creator": excel_meta["creator"],
                "last_modified_by": excel_meta["last_modified_by"],
            })
        elif creator:
            checks.append({
                "check": "different_modifier",
                "severity": "ok",
                "message": f"Created and last modified by same user: '{excel_meta['creator']}'",
            })

        # Check 3: Excel internal dates vs file system dates
        excel_created = excel_meta.get("created")
        excel_modified = excel_meta.get("modified")
        if excel_created and created:
            try:
                excel_dt = datetime.fromisoformat(excel_created)
                fs_dt = datetime.fromisoformat(created)
                diff_hours = abs((excel_dt - fs_dt).total_seconds()) / 3600
                if diff_hours > 24:
                    checks.append({
                        "check": "date_mismatch",
                        "severity": "warning",
                        "message": f"Excel creation date ({excel_created[:10]}) differs from filesystem ({created[:10]}) by {diff_hours / 24:.0f} days",
                        "excel_date": excel_created,
                        "filesystem_date": created,
                    })
                else:
                    checks.append({
                        "check": "date_mismatch",
                        "severity": "ok",
                        "message": "Excel and filesystem creation dates match",
                    })
            except (ValueError, TypeError):
                pass

        # Check 4: Authoring software
        creator_sw = excel_meta.get("creator") or ""
        if any(kw in creator_sw.lower() for kw in ("google", "libreoffice", "wps", "numbers")):
            checks.append({
                "check": "authoring_software",
                "severity": "info",
                "message": f"File created with '{creator_sw}' — not original bank export software",
                "software": creator_sw,
            })

    # Check 5: File size sanity
    size = meta.get("file_size_bytes", 0)
    if size < 1024:
        checks.append({
            "check": "file_size",
            "severity": "warning",
            "message": f"File is unusually small ({_format_size(size)}) — may be empty or corrupted",
        })
    elif size > 50 * 1024 * 1024:
        checks.append({
            "check": "file_size",
            "severity": "info",
            "message": f"Large file ({_format_size(size)}) — may take longer to process",
        })
    else:
        checks.append({
            "check": "file_size",
            "severity": "ok",
            "message": f"File size normal ({_format_size(size)})",
        })

    return checks
